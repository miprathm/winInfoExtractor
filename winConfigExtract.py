import sys, os
import re
import openpyxl
import sys, traceback

pathname = os.path.abspath(sys.argv[0])
if len(sys.argv) > 1 :
	pathname = os.path.abspath(sys.argv[1])
print(pathname)

"""

"""
def exception(row_index,file):
	sheet.cell(row=row_index,column=15).value = file
	sheet.cell(row=row_index,column=16).value = "Parsed Error"
	missing_file.write("\n "+file)
	print("\n** Parsing Problem : "+file+" \n Error ")
	traceback.print_exc()

wb = openpyxl.Workbook()
wb.save('info.xlsx')
sheet = wb.get_active_sheet()
missing_file = open("missing",'w')
row_index = 1
for file in os.listdir(pathname):
	if "programs" in file :
		continue
	"""if row_index is 2:
		break
	"""
	script_file = open(os.path.join(pathname,file),'r')
	script = script_file.read()
	#print(script)
	print("\n Processing : "+file)
	script_file.close()
	# Host Name
	#print(os.path.splitext(file)[0])
	# os.path.splitext("path_to_file")[0]
	host_name_finder  = re.compile(r'Host Name\:\s+(.+)')
	ip_finder = re.compile(''' IPv4\sAddress.+\:\s(.+) ''',re.X)
	os_version_finder = re.compile(r'OS Version\:\s+(.+)')
	processor_finder = re.compile(r'Name\s+(.+)')
	system_mfg_finder = re.compile(r'System Manufacturer\:\s+(.+)')
	system_model_finder = re.compile(r'System Model\:\s+(.+)')
	system_type_finder = re.compile(r'System Type\:\s+(.+)')
	total_physical_mem_finder = re.compile(r'Total Physical Memory\:\s+(.+)')
	HDD_size_finder = re.compile(r'Disk 0\s+Online\s+(\d{1,4}\s[GMK]?B)\s+(\d{1,4}\s[GMK]?B)')
	domain_finder = re.compile(r'Domain\:\s+(.+)')
	try:
		host_name = host_name_finder.search(script)
		sheet.cell(row=row_index,column=1).value = host_name.group(1)
	except Exception as e: exception(row_index,file)
	try:	
		ip = ip_finder.search(script)
		sheet.cell(row=row_index,column=2).value = ip.group(1)
	except Exception as e: exception(row_index,file)
	try:	
		os_version = os_version_finder.search(script)
		sheet.cell(row=row_index,column=3).value = os_version.group(1)
	except Exception as e: exception(row_index,file)
	try:	
		processor = processor_finder.search(script)
		sheet.cell(row=row_index,column=4).value = processor.group(1) 
	except Exception as e: exception(row_index,file)
	try:	
		system_mfg = system_mfg_finder.search(script)
		sheet.cell(row=row_index,column=5).value = system_mfg.group(1)
	except Exception as e: exception(row_index,file)
	try:	
		system_model = system_model_finder.search(script)
		sheet.cell(row=row_index,column=6).value = system_model.group(1)
	except Exception as e: exception(row_index,file)
	try:	
		system_type = system_type_finder.search(script)
		sheet.cell(row=row_index,column=7).value = system_type.group(1)
	except Exception as e: exception(row_index,file)
	try:	
		total_physical_mem = total_physical_mem_finder.search(script)
		sheet.cell(row=row_index,column=8).value = total_physical_mem.group(1)
	except Exception as e: exception(row_index,file)
	try:	
		HDD_size = HDD_size_finder.search(script)
		sheet.cell(row=row_index,column=9).value = HDD_size.group(1)
		#HDD_size.group(1)
		sheet.cell(row=row_index,column=10).value = HDD_size.group(2)
		#free.group(1)
	except Exception as e: exception(row_index,file)
	try:	
		domain = domain_finder.search(script)
		sheet.cell(row=row_index,column=11).value = domain.group(1)
	except Exception as e: exception(row_index,file)
	try:	
		program_file = open(os.path.join(pathname,os.path.splitext(file)[0]+"_programs.txt"))
		
		program_details = program_file.read()
		sheet.cell(row=row_index,column=12).value = str(program_details)
		
		program_file.close()
		#sheet.cell(row=row_index,column=12).value = 
		#ip = os_version_finder.search(script)
		#sheet.cell(row=row_index,column=1).value = ip.group(0)
		
		#print(host_name.group(0))
		#print(ip.group(0))
	except Exception as e: exception(row_index,file)
	row_index += 1	
wb.save('info.xlsx')	
missing_file.close()

